from Utility.CryptListModule import *
from Model.Database.DatabaseModel import *
from Model.Record.RecordTableModel import *
from Excel.ExcelFileNameConfig import *
from openpyxl import *
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string


class ExcelDBInfo:
    """
    이전 버전의 Excel Database에서 파일을 가져오기 위한 정보 클래스
    * FieldInfo: DatabaseFieldModelConfig의 field를 key로 가지고, 그 필드에 해당하는 데이터 딕셔너리를 value로 가진다.
        title: Excel Database에서의 field name. 실질적으로 사용하지 않는다.
        column: Excel Database에서 key에 해당하는 값(DatabaseModel에 입력되어야 하는 값)을 가지고 있는 column 값.
    * StartRow: 입력받을 데이터가 존재하는 첫 행. EndRow는 importExcelDatabase 함수에서 자체적으로 계산함.
                - 조건: 데이터와 데이터 사이에 빈 행이 있으면 그 후는 인식하지 못한다.
    * excelTextFilter(): 읽은 cell값을 받아서 str로 변환해주는 메소드.
    """
    FieldInfo = {
        '성명': {
            'title': '성명',
            'column': 4
        },
        '생년월일': {
            'title': '생년월일',
            'column': 5
        },
        '차량번호':{
            'title': '차량번호',
            'column': 6
        },
        '소속': {
            'title': '소속',
            'column': 7
        },
        '방문목적': {
            'title': '목적',
            'column': 8
        },
        '비고': {
            'title': '비고(저장만됨)',
            'column': 9
        },
        '최초출입날짜': {
            'title': '최근 들어온 시간',
            'column': 10
        },
        '최근출입날짜': {
            'title': '최근 들어온 시간',
            'column': 10
        }
    }
    StartRow = 16

    @classmethod
    def excelValueToString(cls, property) -> str:
        """
        :param property: openpyxl을 통해 입력받은 cell값
        :return: 변환된 str값
        """
        if isinstance(property, datetime):
            return property.strftime('%Y-%m-%d')
        elif property is None:
            return ''
        else:
            return str(property)


class ExcelReportInfo:
    """
    마감 파일을 작성할 때 필요한 정보 클래스 (베이스가 되는 마감파일명은 ExcelFileNameConfig에 있음)
    * FieldColumnDict: RecordFieldModelConfig의 field를 key로 가지고, property를 작성할 column 값을 value로 가짐
    * StartRow: 데이터 작성을 시작할 행 값
    * DeliverySpanWidth: 인수인계 전달사항 항목에서 가로 길이
    * InfoCellAddress: 정보전달 cell의 주소를 가지고 있는 딕셔너리.
                        date - report 작성 날짜, 포맷은 yyMMdd
                        location - report의 장소-근무지, 포맷은 '장소{space}근무지'
    * CellBorder: 데이터가 작성되는 셀에 적용되는 테두리값.
    * setFieldsColumn(): field-column 딕셔너리를 입력받아 그 값을 FieldColumnDict에 적용시킴
    """
    FieldColumnDict = {
        '출입증번호': 1,
        '성명': 2,
        '생년월일': 3,
        '비고': 4,
        '차량번호': 5,
        '소속': 6,
        '방문목적': 7,
        '반출입물품명': 8,
        '반입/반출량': 9,
        '들어오다시간': 10,
        '들어오다근무자': 11,
        '나가다시간': 12,
        '나가다근무자': 13
    }
    StartRow = 7
    DeliverySpanWidth = 6
    InfoCellAddress = {
        'date': 'A1',
        'location': 'D1',
        'takeover_start': 'O5'
    }
    CellBorder = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

    # @classmethod
    # def setFieldsColumn(cls, field_column_dict: Dict[str, int]) -> bool:
    #     """
    #     :param field_column_dict
    #     :return: **
    #         field_column_dict의 입력 결과 FieldColumnDict에 column 중복값이 발생하였다면,
    #         변경을 취소하고 변경 실패의 False를 반환함
    #         중복값이 발생하지 않는다면 변경을 유지하고 True를 반환함.
    #     """
    #     original_dict = cls.FieldColumnDict.copy()
    #     for field, column in field_column_dict.items():
    #         if field in cls.FieldColumnDict.keys():
    #             cls.FieldColumnDict[field] = column
    #     column_list = list(field_column_dict.values())
    #     if len(column_list) != len(set(column_list)):
    #         cls.FieldColumnDict = original_dict.copy()
    #         return False
    #     else:
    #         return True


class ExcelFileModule:
    @classmethod
    def importExcelDatabase(cls, location_string: str, file_name: str) -> Optional[DatabaseModel]:
        """
        file_name(ExcelFileNameConfig 참조)에 있는 이전 버전 Excel Database를
        Arisu.exe에서 사용할 수 있는 iml Database Model로 변환함.
        만들어지는 database의 location은 location string 인자를 따름
        """
        print('ExcelModule setting')
        db_file = load_workbook(file_name)
        db_sheet = db_file[ExcelFileNameConfig.getExcelDatabaseSheetName()]  # Sheet1
        info = ExcelDBInfo
        start_row, end_row = info.StartRow, db_sheet.max_row
        db_model = DatabaseModel(location_string)
        if db_model.hasFile():
            return None

        print('Start loading')
        db_model.blockSignals(True)
        for row_iter in range(start_row, end_row + 1):
            property_dict = {field: '' for field in DatabaseFieldModelConfig.getFieldList()}
            for field_iter, info_iter in info.FieldInfo.items():
                field_column = info_iter['column']
                property = info.excelValueToString(db_sheet.cell(row=row_iter, column=field_column).value)
                property_dict[field_iter] = property
            db_model.addData(VisitorModel(property_dict))
            print(row_iter, 'row data added')
        db_model.blockSignals(False)
        db_file.close()

        print('Finish loading')
        db_model.save()
        return db_model

    @classmethod
    def exportExcelRecord(cls, location_string: str, date_string: str) -> str:
        """
        인자로 받은 정보에 해당하는 RecordTableModel을 엑셀 마감 파일로 바꿈.
        반환하는 값은 생성한 마감 파일의 이름
            1. date와 location 위치에 record의 날짜와 record의 location을 삽입함
            2. recordTable의 모든 데이터를 info에 맞춰서 작성함
            3. report 파일을 생성하고 그 이름을 반환
        """
        print('ExcelModule setting')
        record_table_model = RecordTableModel(location_string, date_string)   # todo: has file 체크하기
        info = ExcelReportInfo
        wb = load_workbook(ExcelFileNameConfig.getExcelReportSampleName())
        sheet = wb['date']
        # sheet title and date
        sheet.title = date_string
        sheet[info.InfoCellAddress['date']] = datetime.strptime(date_string, '%y%m%d').strftime('%Y-%m-%d')
        # view header
        title_string = str(sheet[info.InfoCellAddress['location']].value)
        sheet[info.InfoCellAddress['location']] = title_string.replace('location', location_string)
        delivery_column, delivery_row = coordinate_from_string(info.InfoCellAddress['takeover_start'])
        delivery_column = column_index_from_string(delivery_column)

        print('Start writing')
        row_iter = info.StartRow
        for data_iter in record_table_model.getDataList():
            if data_iter.isTakeoverRecord():
                col_list = info.FieldColumnDict.values()
                min_col, max_col = min(col_list), max(col_list)
                takeover_string, delivery_string = data_iter.getProperty('인수인계'), data_iter.getProperty('비고')
                sheet.merge_cells(start_row=row_iter, start_column=min_col, end_row=row_iter, end_column=max_col)
                takeover_cell = sheet.cell(row=row_iter, column=min_col)
                takeover_cell.value = data_iter.getProperty('인수인계')
                takeover_cell.alignment = Alignment(horizontal='center', vertical='center')
                takeover_cell.font = Font(bold=True)
                for col_iter in range(min_col, max_col+1):
                    sheet.cell(row=row_iter, column=col_iter).border = info.CellBorder

                # 인수인계 전달사항 제목
                delivery_end_column = delivery_column + info.DeliverySpanWidth - 1
                sheet.merge_cells(start_row=delivery_row, start_column=delivery_column,
                                  end_row=delivery_row, end_column=delivery_end_column)
                delivery_title_cell = sheet.cell(row=delivery_row, column=delivery_column)
                delivery_title_cell.value = takeover_string
                delivery_title_cell.alignment = Alignment(horizontal='center', vertical='center')
                delivery_title_cell.font = Font(bold=True)
                for col_iter in range(delivery_column, delivery_end_column + 1):
                    sheet.cell(row=delivery_row, column=col_iter).border = info.CellBorder

                # 인수인계 전달사항 내용
                newline_count = delivery_string.count('\n')
                sheet.merge_cells(start_row=delivery_row + 1, start_column=delivery_column,
                                  end_row=delivery_row + 1 + newline_count, end_column=delivery_end_column)
                delivery_content_cell = sheet.cell(row=delivery_row + 1, column=delivery_column)
                indented_delivery_string = ' ' + delivery_string.replace('\n', '\n ')
                delivery_content_cell.value = indented_delivery_string
                delivery_content_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                for delivery_row_iter in range(delivery_row + 1, delivery_row + 1 + newline_count + 1):
                    for col_iter in range(delivery_column, delivery_end_column + 1):
                        sheet.cell(row=delivery_row_iter, column=col_iter).border = info.CellBorder
                delivery_row = delivery_row + 1 + newline_count + 3  # title(1) / lineCount(newline+1) / spaceHeight(2)
                print('Takeover added:', takeover_string)

            else:
                for field_iter, col_iter in info.FieldColumnDict.items():
                    data_cell = sheet.cell(row=row_iter, column=col_iter)
                    data_cell.value = data_iter.getProperty(field_iter)
                    data_cell.alignment = Alignment(horizontal='center', vertical='center')
                    data_cell.border = info.CellBorder
            # 하단은 inserted와 finished 데이터만 삽입하는 코드
            # elif data_iter.getState() in [RecordModel.State.Inserted, RecordModel.State.Finished]:
            # pass
            # else:
            #     continue
            print(row_iter, 'row data wrote')
            row_iter += 1

        print('Finish writing')
        directory, file_name = FilePathConfig.getReportPath(location_string, date_string)
        if directory:
            file_path = directory + '\\' + file_name
        else:
            file_path = file_name
        wb.save(file_path)
        wb.close()
        print(file_path, 'saved')
        return file_name

