from Model.Table.Database.DatabaseTableModel import *
from Model.Table.Record.RecordTableModel import *
from Utility.Module.Excel.ExcelModuleInformation import *
from datetime import date
from openpyxl import *
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string


"""
ExcelModule
전역 클래스
1. RecordTableModel <-> 엑셀 기록부(마감) 파일
2. DatabaseTableModel <-> 엑셀 데이터베이스 파일
각각의 전환에는 information.json 파일을 사용함
"""


class ExcelModule:
    """
    advanced property
    * cellBorder
    * deliverySpanWidth (record)
    """
    @classmethod
    def cellBorder(cls) -> Border:
        return Border(left=Side(style='thin'), right=Side(style='thin'),
                      top=Side(style='thin'), bottom=Side(style='thin'))

    @classmethod
    def deliverySpanWidth(cls) -> int:
        return 6

    """
    method
    * convertDatabaseFromExcel, convertRecordFromExcel
    * convertDatabaseToExcel, convertRecordToExcel
    """
    @classmethod
    def convertDatabaseFromExcel(cls, excel_file_path: str) -> None:
        print('ExcelModule setting: Database Import from Excel /', excel_file_path)
        import_database_path = cls.__readDatabaseFromExcel(excel_file_path)
        print(f'ExcelModule Finished - Read At [{import_database_path}]')

    @classmethod
    def convertRecordFromExcel(cls, excel_file_path: str) -> None:
        print('ExcelModule setting: Record Import from Excel /', excel_file_path)
        import_record_path = cls.__readRecordFromExcel(excel_file_path)
        if import_record_path:
            print(f'ExcelModule Finished - Read At [{import_record_path}]')
        else:
            print('Error')

    @classmethod
    def convertDatabaseToExcel(cls, database_file_path: str) -> None:
        [location] = ConfigModule.FilePath.fileNameToData(FileType.DatabaseTable, database_file_path)
        table_model = DatabaseTableModel(location)
        print('ExcelModule setting: Database Export to Excel /', table_model.filePath())
        export_file_path = cls.__writeDatabaseExcel(table_model)
        if export_file_path:
            print(f'ExcelModule Finished - Write At [{export_file_path}]')
        else:
            print('Error')

    @classmethod
    def convertRecordToExcel(cls, record_file_path: str) -> None:
        table_model = RecordTableModel(record_file_path)
        print('ExcelModule setting: Record Export to Excel /', table_model.filePath())
        export_file_path = cls.__writeRecordExcel(table_model)
        if export_file_path:
            print(f'ExcelModule Finished - Write At [{export_file_path}]')
        else:
            print('Error')

    """
    private method
    * __readDatabaseFromExcel, __readRecordFromExcel
    * __writeDatabaseExcel, __writeRecordExcel
    """
    @classmethod
    def __readDatabaseFromExcel(cls, excel_file_path: str) -> str:
        Info: Dict[str, Any] = ExcelModuleInformation.excelInformation(ExcelModuleInformation.MethodType.DatabaseFromExcel)
        InfoKeyType = ExcelModuleInformation.InfoKey
        print('read start')

        # 엑셀 열기
        excel_file = load_workbook(excel_file_path)
        sheet = excel_file.worksheets[0]  # 첫번째 시트

        # 필수 정보 작성 및 로딩 (시작을 대문자로) (장소, 날짜, delivery 작성 시작 위치 등등)
        def findInfoFromCellText(text: str) -> str:
            if '(' in text:
                left_paren, right_paren = text.find('('), text.find(')')
                return text[left_paren + 1:right_paren].strip()
            else:
                return text
        location = findInfoFromCellText(str(sheet[Info[InfoKeyType.Location]].value))
        Field_row, Data_start_row = Info[InfoKeyType.FieldRow], Info[InfoKeyType.DataStartRow]
        Field_column_dict = Info[InfoKeyType.FieldColumn]
        print('Sample and Data Load Success')
        print(f'Location: {location}')

        # 작성할 데이터베이스 파일 로딩
        if os.path.isfile(ConfigModule.FilePath.databaseTableFilePath(location)):
            ErrorLogger.reportError(f'이미 [{location}]에 해당하는 데이터베이스 파일이 존재합니다.\n'
                                    f'{ConfigModule.FilePath.databaseTableFilePath(location)}', NameError)
            return None
        database_model = DatabaseTableModel(location)

        # 데이터 읽기 시작
        print('Start loading')
        database_model.blockSignals(True)
        row_iter = Data_start_row
        while True:
            data_dict_iter = {}
            for field_name_iter, column_iter in Field_column_dict.items():
                data_iter = sheet.cell(row=row_iter, column=column_iter).value
                if isinstance(data_iter, datetime):
                    field_model = ConfigModule.TableField.fieldModel(field_name_iter)
                    if field_model.globalOption(TableFieldOption.Global.IsTime):
                        data_iter = data_iter.strftime('%H:%M')
                    elif field_model.globalOption(TableFieldOption.Global.IsDate):
                        data_iter = data_iter.strftime('%Y-%m-%d')
                    else:
                        data_iter = str(data_iter)
                elif data_iter is None:
                    data_iter = ''
                else:
                    data_iter = str(data_iter)
                data_dict_iter[field_name_iter] = data_iter
            if not any(data_dict_iter.values()):
                break
            database_model.addItem(data_dict_iter)
            print(row_iter, 'row data added')
            row_iter += 1

        print('Finish loading')
        database_model.blockSignals(False)
        excel_file.close()
        database_model.save()
        print('File Save')
        return database_model.filePath()

    @classmethod
    def __readRecordFromExcel(cls, excel_file_path: str) -> str:
        Info: Dict[str, Any] = ExcelModuleInformation.excelInformation(ExcelModuleInformation.MethodType.RecordFromExcel)
        InfoKeyType = ExcelModuleInformation.InfoKey
        print('read start')

        # 엑셀 열기
        excel_file = load_workbook(excel_file_path)
        sheet = excel_file.worksheets[0]  # 첫번째 시트

        # 필수 정보 작성 및 로딩 (시작을 대문자로) (장소, 날짜, delivery 작성 시작 위치 등등)
        def findInfoFromCellText(text: str) -> str:
            if '(' in text:
                left_paren, right_paren = text.find('('), text.find(')')
                return text[left_paren + 1:right_paren].strip()
            else:
                return text
        location = findInfoFromCellText(str(sheet[Info[InfoKeyType.Location]].value))
        date = findInfoFromCellText(str(sheet[Info[InfoKeyType.Date]].value))
        convert_date = datetime.strptime(date, '%Y-%m-%d').strftime('%y%m%d')
        Field_row, Data_start_row = Info[InfoKeyType.FieldRow], Info[InfoKeyType.DataStartRow]
        Field_column_dict = Info[InfoKeyType.FieldColumn]
        Delivery_start_column, Delivery_start_row = coordinate_from_string(Info[InfoKeyType.TakeoverStartCell])
        Delivery_start_column = column_index_from_string(Delivery_start_column)
        print('Sample and Data Load Success')
        print(f'Location: {location} / Date: {convert_date}')

        # 작성할 레코드 파일 로딩
        if os.path.isfile(ConfigModule.FilePath.recordTableFilePath(location, convert_date)):
            ErrorLogger.reportError(f'이미 [{location} / {convert_date}]에 해당하는 기록부 파일이 존재합니다.\n'
                                    f'{ConfigModule.FilePath.recordTableFilePath(location, convert_date)}', NameError)
            return None
        record_model = RecordTableModel(location, convert_date)

        # 데이터 읽기 시작
        print('Start loading')
        record_model.blockSignals(True)
        row_iter, delivery_row_iter = Data_start_row, Delivery_start_row
        while True:
            data_dict_iter = {}
            # 인수인계 처리
            if type(sheet.cell(row=row_iter, column=2)).__name__ == 'MergedCell':
                # read takeover string
                takeover_string = str(sheet.cell(row=row_iter, column=1).value)
                time_string = takeover_string[0:5]  # todo 그냥 대충 5개 슬라이싱 해서 사용. 규격 달라지면 호환 안됨
                print(takeover_string, time_string)  # todo 삭제 필요함
                data_dict_iter[TableFieldOption.Necessary.TAKEOVER] = takeover_string
                data_dict_iter[TableFieldOption.Necessary.IN_TIME] = time_string
                data_dict_iter[TableFieldOption.Necessary.OUT_TIME] = time_string
                # read delivery string
                delivery_string = str(sheet.cell(row=delivery_row_iter+1, column=Delivery_start_column).value)
                cell_temp_iter = sheet.cell(row=delivery_row_iter, column=Delivery_start_column)
                while cell_temp_iter.value or type(cell_temp_iter).__name__ == 'MergedCell':
                    delivery_row_iter += 1
                    cell_temp_iter = sheet.cell(row=delivery_row_iter, column=Delivery_start_column)
                delivery_row_iter += 2
                data_dict_iter[TableFieldOption.Necessary.RECORD_ID] = delivery_string
            # 다른 데이터 처리
            else:
                for field_name_iter, column_iter in Field_column_dict.items():
                    data_iter = sheet.cell(row=row_iter, column=column_iter).value
                    if isinstance(data_iter, datetime):
                        field_model = ConfigModule.TableField.fieldModel(field_name_iter)
                        if field_model.globalOption(TableFieldOption.Global.IsTime):
                            data_iter = data_iter.strftime('%H:%M')
                        elif field_model.globalOption(TableFieldOption.Global.IsDate):
                            data_iter = data_iter.strftime('%Y-%m-%d')
                        else:
                            data_iter = str(data_iter)
                    elif data_iter is None:
                        data_iter = ''
                    else:
                        data_iter = str(data_iter)
                    data_dict_iter[field_name_iter] = data_iter
            if not any(data_dict_iter.values()):
                break
            record_model.addItem(data_dict_iter)
            print(row_iter, 'row data added')
            row_iter += 1

        print('Finish loading')
        record_model.blockSignals(False)
        excel_file.close()
        record_model.save()
        print('File Save')
        return record_model.filePath()

    @classmethod
    def __writeDatabaseExcel(cls, table_model: DatabaseTableModel) -> str:
        Info: Dict[str, Any] = ExcelModuleInformation.excelInformation(ExcelModuleInformation.MethodType.DatabaseToExcel)
        InfoKeyType = ExcelModuleInformation.InfoKey
        print('write start')

        # 기본 필수값 로딩
        location = table_model.location()

        # 엑셀 열기
        wb = load_workbook(DefaultFilePath.ExcelDatabaseSample)
        sheet = wb['DB']

        # 필수 정보 작성 및 로딩 (시작을 대문자로) (장소, 날짜, delivery 작성 시작 위치 등등)
        location_address = Info[InfoKeyType.Location]
        sheet[location_address] = str(sheet[location_address].value).replace('location', location)
        Field_row, Data_start_row = Info[InfoKeyType.FieldRow], Info[InfoKeyType.DataStartRow]
        Field_column_dict = Info[InfoKeyType.FieldColumn]
        print('Sample and Data Load Success')

        # 데이터 작성 시작
        print('Start writing')
        row_iter = Data_start_row
        for item_iter in table_model.itemList():
            # 데이터 작성시
            for field_iter, col_iter in Field_column_dict.items():
                data_cell = sheet.cell(row=row_iter, column=col_iter)
                data_cell.value = item_iter.fieldData(field_iter)
                data_cell.alignment = Alignment(horizontal='center', vertical='center')
                data_cell.border = cls.cellBorder()
            print(row_iter, 'row data wrote')
            row_iter += 1

        print('Finish writing')
        excel_database_path = ConfigModule.FilePath.excelDatabaseFilePath(location)
        wb.save(excel_database_path)
        print('File Save')
        wb.close()
        return excel_database_path

    @classmethod
    def __writeRecordExcel(cls, table_model: RecordTableModel) -> str:
        Info: Dict[str, Any] = ExcelModuleInformation.excelInformation(ExcelModuleInformation.MethodType.RecordToExcel)
        InfoKeyType = ExcelModuleInformation.InfoKey
        print('write start')

        # 기본 필수값 로딩
        location, date = table_model.location(), table_model.date()
        convert_date = datetime.strptime(date, '%y%m%d').strftime('%Y-%m-%d')

        # 엑셀 열기
        wb = load_workbook(DefaultFilePath.ExcelRecordSample)
        sheet = wb['date']
        sheet.title = date

        # 필수 정보 작성 및 로딩 (시작을 대문자로) (장소, 날짜, delivery 작성 시작 위치 등등)
        location_address = Info[InfoKeyType.Location]
        sheet[location_address] = str(sheet[location_address].value).replace('location', location)
        date_address = Info[InfoKeyType.Date]
        sheet[date_address] = str(sheet[date_address].value).replace('date', convert_date)
        Field_row, Data_start_row = Info[InfoKeyType.FieldRow], Info[InfoKeyType.DataStartRow]
        Field_column_dict = Info[InfoKeyType.FieldColumn]
        Delivery_start_column, Delivery_start_row = coordinate_from_string(Info[InfoKeyType.TakeoverStartCell])
        Delivery_start_column = column_index_from_string(Delivery_start_column)
        print('Sample and Data Load Success')

        # 데이터 작성 시작
        print('Start writing')
        row_iter, delivery_row_iter = Data_start_row, Delivery_start_row
        for item_iter in table_model.itemList():
            # 인수인계 데이터 작성시
            if item_iter.state() == RecordModel.State.Takeover:
                col_list = Field_column_dict.values()
                min_col, max_col = min(col_list), max(col_list)
                takeover_string = item_iter.fieldData(TableFieldOption.Necessary.TAKEOVER)
                delivery_string = item_iter.fieldData(TableFieldOption.Necessary.RECORD_ID)
                sheet.merge_cells(start_row=row_iter, start_column=min_col, end_row=row_iter, end_column=max_col)
                takeover_cell = sheet.cell(row=row_iter, column=min_col)
                takeover_cell.value = takeover_string
                takeover_cell.alignment = Alignment(horizontal='center', vertical='center')
                takeover_cell.font = Font(bold=True)
                for col_iter in range(min_col, max_col + 1):
                    sheet.cell(row=row_iter, column=col_iter).border = cls.cellBorder()

                # 인수인계 전달사항 제목
                delivery_end_column = Delivery_start_column + cls.deliverySpanWidth() - 1
                sheet.merge_cells(start_row=delivery_row_iter, start_column=Delivery_start_column,
                                  end_row=delivery_row_iter, end_column=delivery_end_column)
                delivery_title_cell = sheet.cell(row=delivery_row_iter, column=Delivery_start_column)
                delivery_title_cell.value = takeover_string
                delivery_title_cell.alignment = Alignment(horizontal='center', vertical='center')
                delivery_title_cell.font = Font(bold=True)
                for col_iter in range(Delivery_start_column, delivery_end_column + 1):
                    sheet.cell(row=delivery_row_iter, column=col_iter).border = cls.cellBorder()

                # 인수인계 전달사항 내용
                newline_count = delivery_string.count('\n')
                sheet.merge_cells(start_row=delivery_row_iter + 1, start_column=Delivery_start_column,
                                  end_row=delivery_row_iter + 1 + newline_count, end_column=delivery_end_column)
                delivery_content_cell = sheet.cell(row=delivery_row_iter + 1, column=Delivery_start_column)
                indented_delivery_string = ' ' + delivery_string.replace('\n', '\n ')
                delivery_content_cell.value = indented_delivery_string
                delivery_content_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                for write_row_iter in range(delivery_row_iter + 1, delivery_row_iter + 1 + newline_count + 1):
                    for col_iter in range(Delivery_start_column, delivery_end_column + 1):
                        sheet.cell(row=write_row_iter, column=col_iter).border = cls.cellBorder()
                delivery_row_iter = delivery_row_iter + 1 + newline_count + 3  # title(1) / lineCount(newline+1) / spaceHeight(2)
                print('Takeover added:', takeover_string)

            # 그 외 데이터 작성시
            else:
                for field_iter, col_iter in Field_column_dict.items():
                    data_cell = sheet.cell(row=row_iter, column=col_iter)
                    data_cell.value = item_iter.fieldData(field_iter)
                    data_cell.alignment = Alignment(horizontal='center', vertical='center')
                    data_cell.border = cls.cellBorder()
            print(row_iter, 'row data wrote')
            row_iter += 1

        print('Finish writing')
        excel_record_path = ConfigModule.FilePath.excelRecordFilePath(location, date)
        wb.save(excel_record_path)
        print('File Save')
        wb.close()
        return excel_record_path
